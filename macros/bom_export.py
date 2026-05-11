"""
BOM Export Macro
----------------
Simulates extracting a Bill of Materials from an assembly tree
and exporting it to Excel.

In real NX/CATIA: replace MOCK_ASSEMBLY with NX Open API calls
to traverse the actual part tree and read attributes.
"""

import base64
from io import BytesIO
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

from pathlib import Path
from datetime import datetime
from core.logger import logger

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ModuleNotFoundError:
    openpyxl = None
    Alignment = Font = PatternFill = None


# Mock assembly tree — simulates data from NX/CATIA assembly
MOCK_ASSEMBLIES = {
    "ENGINE_BLOCK_v3": [
        {"part_no": "PN-001", "description": "Cylinder Block",     "material": "Cast Iron",      "qty": 1,  "revision": "C", "weight_kg": 45.2},
        {"part_no": "PN-002", "description": "Piston",             "material": "Aluminium Alloy","qty": 6,  "revision": "B", "weight_kg": 0.8},
        {"part_no": "PN-003", "description": "Crankshaft",         "material": "Forged Steel",   "qty": 1,  "revision": "D", "weight_kg": 12.5},
        {"part_no": "PN-004", "description": "Connecting Rod",     "material": "Alloy Steel",    "qty": 6,  "revision": "A", "weight_kg": 0.6},
        {"part_no": "PN-005", "description": "Camshaft",           "material": "Chilled Iron",   "qty": 2,  "revision": "B", "weight_kg": 3.2},
        {"part_no": "PN-006", "description": "Valve (Intake)",     "material": "Stainless Steel","qty": 12, "revision": "A", "weight_kg": 0.1},
        {"part_no": "PN-007", "description": "Valve (Exhaust)",    "material": "Nimonic Alloy",  "qty": 12, "revision": "A", "weight_kg": 0.1},
        {"part_no": "PN-008", "description": "Oil Pump",           "material": "Aluminium",      "qty": 1,  "revision": "C", "weight_kg": 1.4},
        {"part_no": "PN-009", "description": "Timing Chain",       "material": "Carbon Steel",   "qty": 1,  "revision": "B", "weight_kg": 0.5},
        {"part_no": "PN-010", "description": "Head Gasket",        "material": "Multi-layer Steel","qty": 1, "revision": "E", "weight_kg": 0.3},
    ],
    "GEARBOX_ASSY": [
        {"part_no": "GB-001", "description": "Gearbox Housing",    "material": "Aluminium Die Cast","qty": 1, "revision": "B", "weight_kg": 8.5},
        {"part_no": "GB-002", "description": "Input Shaft",        "material": "Case Hardened Steel","qty": 1,"revision": "A", "weight_kg": 2.1},
        {"part_no": "GB-003", "description": "Output Shaft",       "material": "Case Hardened Steel","qty": 1,"revision": "A", "weight_kg": 3.4},
        {"part_no": "GB-004", "description": "Synchromesh Ring",   "material": "Brass",           "qty": 6, "revision": "C", "weight_kg": 0.2},
        {"part_no": "GB-005", "description": "Gear Selector Fork", "material": "Steel",           "qty": 3, "revision": "A", "weight_kg": 0.4},
    ]
}


def _column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def _xlsx_cell(row_idx: int, col_idx: int, value) -> str:
    ref = f"{_column_name(col_idx)}{row_idx}"
    if isinstance(value, (int, float)):
        return f'<c r="{ref}"><v>{value}</v></c>'
    return f'<c r="{ref}" t="inlineStr"><is><t>{escape(str(value))}</t></is></c>'


def _build_basic_xlsx(headers: list[str], rows: list[list]) -> bytes:
    sheet_rows = []
    for row_idx, row in enumerate([headers] + rows, 1):
        cells = "".join(_xlsx_cell(row_idx, col_idx, value) for col_idx, value in enumerate(row, 1))
        sheet_rows.append(f'<row r="{row_idx}">{cells}</row>')

    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{''.join(sheet_rows)}</sheetData>"
        "</worksheet>"
    )

    stream = BytesIO()
    with ZipFile(stream, "w", ZIP_DEFLATED) as archive:
        archive.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
            '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
            "</Types>",
        )
        archive.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
            "</Relationships>",
        )
        archive.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Bill of Materials" sheetId="1" r:id="rId1"/></sheets>'
            "</workbook>",
        )
        archive.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
            "</Relationships>",
        )
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return stream.getvalue()


def _build_report_bytes(headers: list[str], rows: list[list]) -> bytes:
    if openpyxl is None:
        return _build_basic_xlsx(headers, rows)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bill of Materials"

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    alt_fill = PatternFill("solid", fgColor="EBF0FA")
    for row_idx, row_data in enumerate(rows, 2):
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(horizontal="center" if col != 2 else "left")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 16

    report_stream = BytesIO()
    wb.save(report_stream)
    return report_stream.getvalue()


def run(params: dict) -> dict:
    assembly_name = params.get("assembly_name", "ENGINE_BLOCK_v3")
    output_format = params.get("output_format", "xlsx")
    persist_report = params.get("_persist_report", True)
    include_report = params.get("_include_report", False)

    logger.info(f"BOM Export started for assembly: {assembly_name}")

    parts = MOCK_ASSEMBLIES.get(assembly_name)
    if not parts:
        available = list(MOCK_ASSEMBLIES.keys())
        raise ValueError(f"Assembly '{assembly_name}' not found. Available: {available}")

    filename = f"BOM_{assembly_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = Path("outputs") / filename

    headers = ["Part No", "Description", "Material", "Qty", "Revision", "Weight (kg)", "Total Weight (kg)"]
    rows = []
    table_rows = []
    for part in parts:
        total_weight = round(part["qty"] * part["weight_kg"], 3)
        rows.append([
            part["part_no"], part["description"], part["material"],
            part["qty"], part["revision"], part["weight_kg"], total_weight
        ])
        table_rows.append({
            "part_no": part["part_no"],
            "description": part["description"],
            "material": part["material"],
            "qty": part["qty"],
            "revision": part["revision"],
            "weight_kg": part["weight_kg"],
            "total_weight_kg": total_weight,
        })

    report_bytes = _build_report_bytes(headers, rows)

    if persist_report:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(report_bytes)

    total_parts = len(parts)
    total_weight = round(sum(p["qty"] * p["weight_kg"] for p in parts), 2)
    destination = output_path if persist_report else "API response"
    logger.success(f"BOM exported: {total_parts} parts, {total_weight}kg total → {destination}")

    output = {
        "file": str(output_path) if persist_report else None,
        "assembly": assembly_name,
        "output_format": output_format,
        "total_parts": total_parts,
        "total_weight_kg": total_weight,
        "table": {
            "columns": [
                {"key": "part_no", "label": "Part No"},
                {"key": "description", "label": "Description"},
                {"key": "material", "label": "Material"},
                {"key": "qty", "label": "Qty"},
                {"key": "revision", "label": "Revision"},
                {"key": "weight_kg", "label": "Weight (kg)"},
                {"key": "total_weight_kg", "label": "Total Weight (kg)"},
            ],
            "rows": table_rows,
        },
    }

    if include_report:
        output["report"] = {
            "filename": filename,
            "media_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "encoding": "base64",
            "content": base64.b64encode(report_bytes).decode("ascii"),
        }

    return output
